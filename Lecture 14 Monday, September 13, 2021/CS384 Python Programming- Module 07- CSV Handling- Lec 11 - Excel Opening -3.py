from openpyxl import load_workbook

# wb = Workbook()
# sheet = wb.active
wb = load_workbook(r"appending_values.xlsx")
sheet = wb.active
# data = (
#     (11, 48, 50),
#     (81, 30, 82),
#     (20, 51, 72),
#     (21, 14, 60),
#     (28, 41, 49),
#     (74, 65, 53),
#     ("Peter", 'Andrew', 45.63),
#     (212, 114, 160),
# )
l = ("hi","append","this")
# for i in data:
#     sheet.append(i)
sheet.append(l)
wb.save('appending_values.xlsx')
